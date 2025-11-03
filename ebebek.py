import requests
import pandas as pd
from datetime import datetime
import getpass
from colorama import Fore, Style, init
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import io
from dotenv import load_dotenv

# .env dosyasını yükle
load_dotenv()

# Windows encoding ve terminal düzeltmesi (exe çıktısı için)
if sys.platform == 'win32':
    import codecs
    # EXE modunda stdout.buffer None olabilir, kontrol et
    try:
        if hasattr(sys.stdout, 'buffer') and sys.stdout.buffer is not None:
            sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer)
        if hasattr(sys.stderr, 'buffer') and sys.stderr.buffer is not None:
            sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer)
    except (AttributeError, OSError):
        # EXE modunda veya buffer yoksa, normal stdout kullan
        pass
    # Terminal genişliğini ayarla (sadece konsol modunda)
    try:
        if hasattr(sys.stdout, 'buffer') and sys.stdout.buffer is not None:
            os.system('mode con: cols=80 lines=30')
    except:
        pass

# `colorama`'yı başlatıyoruz
init(autoreset=True)

# Google Sheets'ten şifre çekme fonksiyonu
def get_password_from_sheets():
    # Sheet ID'yi environment variable'dan al
    sheet_id = os.getenv('SHEET_ID')
    if not sheet_id:
        print(Fore.RED + "✗ HATA: SHEET_ID .env dosyasında tanımlı değil!")
        return None
    
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet=Sheet1"
    
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            # CSV verisini oku
            lines = response.text.strip().split('\n')
            if len(lines) > 0:
                # İlk satır (B1 hücresi ikinci sütun)
                first_row = lines[0].split(',')
                if len(first_row) > 1:
                    # Tırnak işaretlerini temizle
                    password = first_row[1].strip('"').strip()
                    return password
        return None
    except Exception as e:
        print(Fore.RED + f"✗ Şifre alınamadı: {e}")
        return None

# Şifreli giriş fonksiyonu
def user_login():
    # .env dosyasından şifre kontrolünün aktif olup olmadığını kontrol et
    # Varsayılan: false (Demo modu aktif)
    password_enabled = os.getenv('PASSWORD_ENABLED', 'false').lower() == 'true'
    
    if not password_enabled:
        print(Fore.GREEN + Style.BRIGHT + """
    ╔════════════════════════════════════════════════════════════╗
    ║                                                            ║
    ║           🏪  EBEBEK BARKOD TAKİP SİSTEMİ  📦              ║
    ║                                                            ║
    ║                  ⚡  DEMO MODU AKTİF  ⚡                    ║
    ║                                                            ║
    ╚════════════════════════════════════════════════════════════╝
    """ + Style.RESET_ALL)
        print(Fore.CYAN + "    ℹ️  Şifre kontrolü devre dışı. Direkt kullanıma geçiliyor...\n")
        return
    
    # Şifre kontrolü aktifse
    # Google Sheets'ten şifreyi çek
    print(Fore.CYAN + "\n" + "═" * 60)
    print(Fore.CYAN + "🔒  Şifre Google Sheets'ten yükleniyor..." + Style.RESET_ALL)
    
    correct_password = get_password_from_sheets()
    
    if correct_password is None:
        print(Fore.RED + "✗ HATA: Şifre yüklenemedi! İnternet bağlantınızı kontrol edin.")
        print(Fore.YELLOW + "⚠  Program sonlandırılıyor...")
        sys.exit()
    
    print(Fore.GREEN + f"✓ Şifre başarıyla yüklendi!\n")
    
    # Modern ASCII Art Başlık
    print(Fore.CYAN + Style.BRIGHT + """
    ╔════════════════════════════════════════════════════════════╗
    ║                                                            ║
    ║           🏪  EBEBEK BARKOD TAKİP SİSTEMİ  📦              ║
    ║                                                            ║
    ║                    🔐  GÜVENLİ GİRİŞ                       ║
    ║                                                            ║
    ╚════════════════════════════════════════════════════════════╝
    """ + Style.RESET_ALL)
    
    attempt = 0
    while True:
        attempt += 1
        entered_password = getpass.getpass(Fore.YELLOW + f"    🔑 Şifre (Deneme {attempt}): " + Style.RESET_ALL)
        
        if entered_password == correct_password:
            print(Fore.GREEN + Style.BRIGHT + "\n    ✓ GİRİŞ BAŞARILI! Hoş Geldiniz! 🎉\n")
            break
        else:
            print(Fore.RED + "    ✗ Yanlış şifre! Lütfen tekrar deneyin.\n")

# Barkodları ve adetleri tutmak için sözlükler
barkodsuz_dict = {}
kayıtsız_list = []
hasarli_dict = {}

# Günün tarihi
today_date = datetime.now().strftime("%d.%m.%Y")

# API URL'si
url = "https://ebebek.wawlabs.com/autocomplete"

# --- Saf işlevler (GUI tarafından da kullanılacak) ---
def validate_barkodsuz(barkod):
    if not barkod or not barkod.isdigit():
        return False, "Barkod yalnızca rakamlardan oluşmalıdır."
    if len(barkod) not in [11, 13]:
        return False, "Barkod 11 veya 13 haneli olmalıdır."
    return True, ""

def validate_kayitsiz(barkod):
    if not barkod or not barkod.startswith("EBHJ"):
        return False, "Barkod 'EBHJ' ile başlamalıdır."
    if len(barkod) != 14:
        return False, "Barkod 14 karakter olmalıdır."
    return True, ""

def format_brand_name(brand_name):
    """Marka adının her kelimesinin baş harfini büyük yapar"""
    if not brand_name:
        return ""
    # Her kelimenin ilk harfini büyük yap
    return " ".join(word.capitalize() for word in str(brand_name).split())

def query_product(barkod, timeout=5):
    try:
        response = requests.get(f"{url}?q={barkod}", timeout=timeout)
        if response.status_code == 200:
            data = response.json()
            products = data.get("products") or []
            return products[0] if products else None
        return None
    except Exception:
        return None

def add_barkodsuz(barkod, product):
    brand = product.get("brand", "Marka bulunamadı") if product else "Bilinmiyor"
    # Marka adını formatla
    brand = format_brand_name(brand)
    product_name = product.get("title", "Ürün adı bulunamadı") if product else "Bilinmiyor"
    if barkod in barkodsuz_dict:
        barkodsuz_dict[barkod]["ADET"] += 1
    else:
        barkodsuz_dict[barkod] = {"TARİH": today_date, "İÇERİK": f"{brand} - {product_name}", "ADET": 1}
    return barkodsuz_dict[barkod]["ADET"]

def add_hasarli(barkod, product):
    brand = product.get("brand", "Marka bulunamadı") if product else "Bilinmiyor"
    # Marka adını formatla
    brand = format_brand_name(brand)
    product_name = product.get("title", "Ürün adı bulunamadı") if product else "Bilinmiyor"
    if barkod in hasarli_dict:
        hasarli_dict[barkod]["ADET"] += 1
    else:
        hasarli_dict[barkod] = {"TARİH": today_date, "İÇERİK": f"{brand} - {product_name}", "ADET": 1}
    return hasarli_dict[barkod]["ADET"]

def add_kayitsiz(barkod):
    if barkod in kayıtsız_list:
        return False
    kayıtsız_list.append(barkod)
    return True

def get_stats():
    barkodsuz_say = len(barkodsuz_dict)
    kayitsiz_say = len(kayıtsız_list)
    hasarli_say = len(hasarli_dict)
    return {
        "barkodsuz": barkodsuz_say,
        "kayitsiz": kayitsiz_say,
        "hasarli": hasarli_say,
        "toplam": barkodsuz_say + kayitsiz_say + hasarli_say,
    }

# Benzersiz dosya adı oluşturma fonksiyonu
def get_unique_filename():
    base_name = f"EBEBEK {today_date}"
    extension = ".xlsx"
    file_name = base_name + extension
    
    if not os.path.exists(file_name):
        return file_name
    
    counter = 1
    while True:
        file_name = f"{base_name} ({counter}){extension}"
        if not os.path.exists(file_name):
            return file_name
        counter += 1

# Modern hoşgeldin ekranı ve ana menü
def welcome_screen():
    os.system('cls' if os.name == 'nt' else 'clear')  # Ekranı temizle
    
    print(Fore.CYAN + Style.BRIGHT + """
    +================================================================+
    |                                                              |
    |          🏪  EBEBEK BARKOD TAKİP SİSTEMİ v2.0  📦            |
    |                                                              |
    |              ⚡ Hızlı • Güvenli • Profesyonel ⚡              |
    |                                                              |
    +================================================================+
    """ + Style.RESET_ALL)
    
    print(Fore.WHITE + "    📅 Tarih        : " + Fore.YELLOW + Style.BRIGHT + today_date)
    print(Fore.WHITE + "    📊 Sistem Durumu: " + Fore.GREEN + Style.BRIGHT + "● Aktif ve Hazır" + Style.RESET_ALL)
    print(Fore.WHITE + "    💾 Veri Tabanı  : " + Fore.GREEN + Style.BRIGHT + "✓ Bağlı\n" + Style.RESET_ALL)
    
    print(Fore.CYAN + Style.BRIGHT + "    +=======================================================+")
    print(Fore.CYAN + "    |" + Fore.WHITE + Style.BRIGHT + "                   📋 İŞLEM MENÜSÜ                   " + Fore.CYAN + "|")
    print(Fore.CYAN + "    +=======================================================+")
    print(Fore.CYAN + "    |                                                       |")
    print(Fore.CYAN + "    |  " + Fore.GREEN + Style.BRIGHT + "【1】" + Style.RESET_ALL + Fore.WHITE + " ➜  🏷️  Barkodsuzlar " + Fore.CYAN + "│" + Fore.WHITE + " 11-13 haneli barkod   " + Fore.CYAN + "|")
    print(Fore.CYAN + "    |  " + Fore.GREEN + Style.BRIGHT + "【2】" + Style.RESET_ALL + Fore.WHITE + " ➜  ❓ Kayıtsızlar  " + Fore.CYAN + "│" + Fore.WHITE + " EBHJ ile başlayan     " + Fore.CYAN + "|")
    print(Fore.CYAN + "    |  " + Fore.GREEN + Style.BRIGHT + "【3】" + Style.RESET_ALL + Fore.WHITE + " ➜  ⚠️  Hasarlılar   " + Fore.CYAN + "│" + Fore.WHITE + " 11-13 haneli barkod   " + Fore.CYAN + "|")
    print(Fore.CYAN + "    |  " + Fore.RED + Style.BRIGHT + "【4】" + Style.RESET_ALL + Fore.WHITE + " ➜  💾 Kaydet & Çık " + Fore.CYAN + "│" + Fore.WHITE + " Excel dosyası oluştur " + Fore.CYAN + "|")
    print(Fore.CYAN + "    |                                                       |")
    print(Fore.CYAN + "    +=======================================================+\n" + Style.RESET_ALL)
    
    # İstatistikler
    toplam = len(barkodsuz_dict) + len(kayıtsız_list) + len(hasarli_dict)
    
    print(Fore.YELLOW + Style.BRIGHT + "    +=======================================================+")
    print(Fore.YELLOW + "    |" + Fore.WHITE + Style.BRIGHT + "               📊 GÜNCEL İSTATİSTİKLER                " + Fore.YELLOW + "|")
    print(Fore.YELLOW + "    +=======================================================+")
    
    # Barkodsuzlar çubuğu
    bar_count = len(barkodsuz_dict)
    bar_visual = "█" * min(bar_count, 20)
    print(Fore.YELLOW + "    |  " + Fore.WHITE + "🏷️  Barkodsuzlar" + Fore.CYAN + f"  : {Fore.GREEN}{bar_count:>3} ürün  " + Fore.GREEN + f"{bar_visual:<20}" + Fore.YELLOW + "  |")
    
    # Kayıtsızlar çubuğu
    kayit_count = len(kayıtsız_list)
    kayit_visual = "█" * min(kayit_count, 20)
    print(Fore.YELLOW + "    |  " + Fore.WHITE + "❓ Kayıtsızlar " + Fore.CYAN + f"  : {Fore.GREEN}{kayit_count:>3} barkod" + Fore.GREEN + f" {kayit_visual:<20}" + Fore.YELLOW + "  |")
    
    # Hasarlılar çubuğu
    hasar_count = len(hasarli_dict)
    hasar_visual = "█" * min(hasar_count, 20)
    print(Fore.YELLOW + "    |  " + Fore.WHITE + "⚠️  Hasarlılar  " + Fore.CYAN + f"  : {Fore.GREEN}{hasar_count:>3} ürün  " + Fore.GREEN + f"{hasar_visual:<20}" + Fore.YELLOW + "  |")
    
    print(Fore.YELLOW + "    +=======================================================+")
    print(Fore.YELLOW + "    |  " + Fore.WHITE + Style.BRIGHT + f"📦 TOPLAM KAYIT : {Fore.CYAN}{toplam:>3} adet                          " + Fore.YELLOW + "|")
    print(Fore.YELLOW + "    +=======================================================+\n" + Style.RESET_ALL)

# Barkodsuzlar işlemi
def barkodsuzlar():
    while True:
        print(Fore.YELLOW + Style.BRIGHT + "\n    +=======================================================+")
        print(Fore.YELLOW + "    |      🏷️  BARKODSUZLAR KATEGORİSİ                 |")
        print(Fore.YELLOW + "    +=======================================================+")
        print(Fore.YELLOW + "    |  " + Fore.WHITE + "📌 Barkod: 11 veya 13 haneli sayı girişi     " + Fore.YELLOW + "|")
        print(Fore.YELLOW + "    |  " + Fore.WHITE + "🔙 Çıkış : Menüye dönmek için 'q' yazın      " + Fore.YELLOW + "|")
        print(Fore.YELLOW + "    +=======================================================+\n" + Style.RESET_ALL)
        
        barkod = input(Fore.WHITE + "    📥 Barkod numarasını girin (menü için 'q'): " + Style.RESET_ALL).strip()

        if barkod.lower() in ['q', 'menu']:
            break

        if (len(barkod) not in [11, 13]) or not barkod.isdigit():
            print(Fore.RED + "    ✗ Hata: Barkod 11 veya 13 haneli rakam olmalıdır!")
            continue

        response = requests.get(f"{url}?q={barkod}")
        if response.status_code == 200:
            data = response.json()

            if "products" in data and len(data["products"]) > 0:
                product = data["products"][0]
                brand = product.get("brand", "Marka bulunamadı")
                product_name = product.get("title", "Ürün adı bulunamadı")

                if barkod in barkodsuz_dict:
                    barkodsuz_dict[barkod]["ADET"] += 1
                    print(Fore.YELLOW + f"    ℹ️  Bu barkod mevcut! Adet artırıldı.")
                else:
                    barkodsuz_dict[barkod] = {"TARİH": today_date, "İÇERİK": f"{brand} - {product_name}", "ADET": 1}

                print(Fore.GREEN + f"    ✓ Ürün: {product_name}")
                print(Fore.GREEN + f"    ✓ Adet: {barkodsuz_dict[barkod]['ADET']}")
            else:
                print(Fore.RED + f"    ✗ {barkod} için ürün bulunamadı.")
        else:
            print(Fore.RED + f"    ✗ API hatası: {response.status_code}")

# Kayıtsızlar işlemi
def kayıtsızlar():
    while True:
        print(Fore.YELLOW + Style.BRIGHT + "\n    +=======================================================+")
        print(Fore.YELLOW + "    |      ❓  KAYITSIZLAR KATEGORİSİ                  |")
        print(Fore.YELLOW + "    +=======================================================+")
        print(Fore.YELLOW + "    |  " + Fore.WHITE + "📌 Format: EBHJ ile başlayan 14 karakter     " + Fore.YELLOW + "|")
        print(Fore.YELLOW + "    |  " + Fore.WHITE + "📝 Örnek : EBHJ1003942527                    " + Fore.YELLOW + "|")
        print(Fore.YELLOW + "    |  " + Fore.WHITE + "🔙 Çıkış : Menüye dönmek için 'q' yazın      " + Fore.YELLOW + "|")
        print(Fore.YELLOW + "    +=======================================================+\n" + Style.RESET_ALL)
        
        barkod = input(Fore.WHITE + "    📥 Barkod numarasını girin (menü için 'q'): " + Style.RESET_ALL).strip()

        if barkod.lower() in ['q', 'menu']:
            break

        if not barkod.startswith("EBHJ") or len(barkod) != 14:
            print(Fore.RED + "    ✗ Hata: Barkod 'EBHJ' ile başlamalı ve 14 karakter olmalı!")
            continue

        if barkod in kayıtsız_list:
            print(Fore.YELLOW + f"    ℹ️  Bu barkod zaten kayıtlı!")
        else:
            kayıtsız_list.append(barkod)
            print(Fore.GREEN + f"    ✓ Barkod kaydedildi: {barkod}")
            print(Fore.CYAN + f"    📊 Toplam: {len(kayıtsız_list)} barkod")

# Hasarlılar işlemi
def hasarlilar():
    while True:
        print(Fore.RED + Style.BRIGHT + "\n    +=======================================================+")
        print(Fore.RED + "    |      ⚠️  HASARLILAR KATEGORİSİ                   |")
        print(Fore.RED + "    +=======================================================+")
        print(Fore.RED + "    |  " + Fore.WHITE + "📌 Barkod: 11 veya 13 haneli sayı girişi     " + Fore.RED + "|")
        print(Fore.RED + "    |  " + Fore.WHITE + "🔙 Çıkış : Menüye dönmek için 'q' yazın      " + Fore.RED + "|")
        print(Fore.RED + "    +=======================================================+\n" + Style.RESET_ALL)
        
        barkod = input(Fore.WHITE + "    📥 Barkod numarasını girin (menü için 'q'): " + Style.RESET_ALL).strip()

        if barkod.lower() in ['q', 'menu']:
            break

        if (len(barkod) not in [11, 13]) or not barkod.isdigit():
            print(Fore.RED + "    ✗ Hata: Barkod 11 veya 13 haneli rakam olmalıdır!")
            continue

        response = requests.get(f"{url}?q={barkod}")
        if response.status_code == 200:
            data = response.json()

            if "products" in data and len(data["products"]) > 0:
                product = data["products"][0]
                brand = product.get("brand", "Marka bulunamadı")
                product_name = product.get("title", "Ürün adı bulunamadı")

                if barkod in hasarli_dict:
                    hasarli_dict[barkod]["ADET"] += 1
                    print(Fore.YELLOW + f"    ℹ️  Bu barkod mevcut! Adet artırıldı.")
                else:
                    hasarli_dict[barkod] = {"TARİH": today_date, "İÇERİK": f"{brand} - {product_name}", "ADET": 1}

                print(Fore.GREEN + f"    ✓ Ürün: {product_name}")
                print(Fore.GREEN + f"    ✓ Adet: {hasarli_dict[barkod]['ADET']}")
            else:
                print(Fore.RED + f"    ✗ {barkod} için ürün bulunamadı.")
        else:
            print(Fore.RED + f"    ✗ API hatası: {response.status_code}")

# Excel dosyasına veri kaydetme
def save_to_excel():
    try:
        if not barkodsuz_dict and not kayıtsız_list and not hasarli_dict:
            print(Fore.YELLOW + "\n    ⚠️  Hiç veri girilmedi! Excel dosyası oluşturulmadı.")
            return

        file_name = get_unique_filename()
        
        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        subheader_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        header_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center")

        # 1. BARKODSUZLAR
        if barkodsuz_dict:
            ws1 = wb.create_sheet("BARKODSUZLAR")
            ws1.merge_cells('A1:C1')
            ws1['A1'] = "EBEBEK BARKODSUZLAR ŞABLONU"
            ws1['A1'].fill = header_fill
            ws1['A1'].font = header_font
            ws1['A1'].alignment = center_align
            
            ws1['A2'] = "TARİH"
            ws1['B2'] = "İÇERİK"
            ws1['C2'] = "ADET"
            for cell in ['A2', 'B2', 'C2']:
                ws1[cell].fill = subheader_fill
                ws1[cell].font = Font(bold=True)
                ws1[cell].alignment = center_align
            
            row = 3
            for barkod, data in barkodsuz_dict.items():
                ws1[f'A{row}'] = data['TARİH']
                ws1[f'B{row}'] = data['İÇERİK']
                ws1[f'C{row}'] = data['ADET']
                ws1[f'C{row}'].alignment = center_align
                row += 1
            
            for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.border = border
            
            ws1.column_dimensions['A'].width = 12
            ws1.column_dimensions['B'].width = 70
            ws1.column_dimensions['C'].width = 10

        # 2. KAYITSIZLAR
        if kayıtsız_list:
            ws2 = wb.create_sheet("KAYITSIZLAR")
            ws2.merge_cells('A1:C1')
            ws2['A1'] = "KAYITSIZLAR"
            ws2['A1'].fill = header_fill
            ws2['A1'].font = header_font
            ws2['A1'].alignment = center_align
            
            ws2['A2'] = "BARKODLAR"
            ws2['B2'] = "BARKODLAR"
            ws2['C2'] = "BARKODLAR"
            for cell in ['A2', 'B2', 'C2']:
                ws2[cell].fill = subheader_fill
                ws2[cell].font = Font(bold=True)
                ws2[cell].alignment = center_align
            
            row = 3
            col = 0
            for barkod in kayıtsız_list:
                col_letter = get_column_letter(col + 1)
                ws2[f'{col_letter}{row}'] = barkod
                ws2[f'{col_letter}{row}'].alignment = center_align
                col += 1
                if col == 3:
                    col = 0
                    row += 1
            
            for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.border = border
            
            ws2.column_dimensions['A'].width = 18
            ws2.column_dimensions['B'].width = 18
            ws2.column_dimensions['C'].width = 18

        # 3. HASARLILAR
        if hasarli_dict:
            ws3 = wb.create_sheet("HASARLILAR")
            ws3.merge_cells('A1:C1')
            ws3['A1'] = "EBEBEK HASARLILAR"
            ws3['A1'].fill = header_fill
            ws3['A1'].font = header_font
            ws3['A1'].alignment = center_align
            
            ws3['A2'] = "TARİH"
            ws3['B2'] = "İÇERİK"
            ws3['C2'] = "ADET"
            for cell in ['A2', 'B2', 'C2']:
                ws3[cell].fill = subheader_fill
                ws3[cell].font = Font(bold=True)
                ws3[cell].alignment = center_align
            
            row = 3
            for barkod, data in hasarli_dict.items():
                ws3[f'A{row}'] = data['TARİH']
                ws3[f'B{row}'] = data['İÇERİK']
                ws3[f'C{row}'] = data['ADET']
                ws3[f'C{row}'].alignment = center_align
                row += 1
            
            for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.border = border
            
            ws3.column_dimensions['A'].width = 12
            ws3.column_dimensions['B'].width = 70
            ws3.column_dimensions['C'].width = 10

        wb.save(file_name)
        print(Fore.GREEN + f"\n    ✓ Dosya kaydedildi: '{file_name}'")
        print(Fore.CYAN + f"\n    📊 ÖZET RAPOR:")
        if barkodsuz_dict:
            print(Fore.CYAN + f"       🏷️  Barkodsuzlar : {len(barkodsuz_dict)} ürün")
        if kayıtsız_list:
            print(Fore.CYAN + f"       ❓  Kayıtsızlar  : {len(kayıtsız_list)} barkod")
        if hasarli_dict:
            print(Fore.CYAN + f"       ⚠️  Hasarlılar   : {len(hasarli_dict)} ürün")
    
    except Exception as e:
        import traceback
        error_msg = f"Excel dosyası oluşturulurken hata: {str(e)}"
        print(Fore.RED + f"\n    ✗ {error_msg}")
        print(traceback.format_exc())
        raise  # Hatayı yukarı fırlat ki Flask yakalayabilsin

# Menü
def menu():
    user_login()
    
    while True:
        welcome_screen()
        choice = input(Fore.WHITE + "    🎯 Seçiminiz: " + Style.RESET_ALL).strip()

        if choice == '1':
            barkodsuzlar()
        elif choice == '2':
            kayıtsızlar()
        elif choice == '3':
            hasarlilar()
        elif choice == '4':
            print(Fore.YELLOW + "\n    📦 Veriler kaydediliyor...")
            save_to_excel()
            print(Fore.GREEN + Style.BRIGHT + "\n    ✓ Program sonlandırıldı. Güle güle! 👋\n")
            break
        else:
            print(Fore.RED + "    ✗ Geçersiz seçim! Lütfen 1-4 arası bir sayı girin.")
            input(Fore.YELLOW + "\n    ⏎ Devam etmek için Enter'a basın...")

# Programı başlat
if __name__ == '__main__':
    menu()